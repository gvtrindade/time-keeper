from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.worksheet.dimensions import ColumnDimension
from tempfile import NamedTemporaryFile

from auths.models import CustomUser
from backend.models import Record

GRAY_COLOR = 'b7e1cd'
YELLOW_COLOR = 'ffff00'
BLUE_COLOR = 'ccffff'

def create_file(year, weeks):
    workbook = Workbook()
    workbook.remove_sheet(workbook.active)
    for w in weeks:
        worksheet = workbook.create_sheet(f'Week - {w}')
        write_file_data(worksheet, year, w)
    
    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        return tmp.read()


def write_file_data(worksheet, year, week):
    global WORKSHEET
    global users_starting_row
    global biggest_name_size
    users_starting_row = 4
    biggest_name_size = 7

    WORKSHEET = worksheet
    week_dates = get_week_dates(year, week)

    write_header(week_dates)
    write_week_days(week_dates)
    write_user_data(year, week)
    set_worksheet_styles()


def write_header(week_dates):

    populate_cell(
        row=1, column=1, 
        value='Record START & FINISH TIME for each shift and the \nlength of each break in HOURS (e.g.0:30)',
        color=GRAY_COLOR,
        font_size=14,
        border=True
    )
    merge_cells(
        start_row=1, start_column=1,
        end_row=1, end_column=10
    )

    first_day = f'{week_dates[0].day} {week_dates[0].strftime("%b ") if week_dates[-1].month != week_dates[0].month else ""}'
    last_day = f'{week_dates[-1].day} {week_dates[-1].strftime("%b")} {week_dates[-1].strftime("%-y")}'
    
    populate_cell(
        row=1, column=11, 
        value=f'TIME SHEET STARTING MONDAY {first_day}- {last_day}', 
        color=YELLOW_COLOR,
        font_size=14,
        border=True
    )
    merge_cells(
        start_row=1, start_column=11,
        end_row=1, end_column=23
    )


def write_week_days(week_dates):
    starting_row = 2
    starting_column = 2
    days = ['Monday', 'Tuesday', 'Wednesday',
            'Thursday', 'Friday', 'Saturday', 'Sunday']

    populate_cell(row=starting_row, column=1, value='Day', color=GRAY_COLOR, border=True)
    merge_cells(
        start_row=starting_row, start_column=1,
        end_row=starting_row + 1, end_column=1
    )

    for day_index in range(7):
        color = None if day_index % 2 != 0 else BLUE_COLOR
        populate_cell(
            starting_row, starting_column, 
            f'{days[day_index]} {ordinal_number(week_dates[day_index].day)}', 
            color=color, border=True
        )
        merge_cells(starting_row, starting_column, starting_row, starting_column + 2)

        populate_cell(starting_row + 1, starting_column, 'Start', color=color, border=True)
        populate_cell(starting_row + 1, starting_column + 1, 'Finish', color=color, border=True)
        populate_cell(starting_row + 1, starting_column + 2, 'Hrs', color=color, border=True)

        starting_column += 3

    populate_cell(starting_row, starting_column, 'Total Hours', border=True)
    merge_cells(starting_row, starting_column, starting_row + 1, starting_column)


def write_user_data(year, week):
    global users_starting_row
    global biggest_name_size

    users = CustomUser.objects.all()
    record = Record()
    for user in users:
        records = record.get_records(user, year, week)

        if user.is_staff and len(records) == 0: return

        populate_cell(users_starting_row, column=1, value=f'{user.first_name} {user.last_name}', color=GRAY_COLOR)

        if (len(f'{user.first_name} {user.last_name}') > biggest_name_size ):
            biggest_name_size = len(f'{user.first_name} {user.last_name}')
        worked_hours = record.calculate_worked_hours(records)
        
        records_starting_row = users_starting_row
        records_starting_column = 2
        offset = 1
        last_saved_record = {}
        for record in records:
            weekday = record.date.weekday()
            assigned_column = records_starting_column + (weekday * 3)
            assigned_row = records_starting_row

            if record.status == 'Wating Approval' or record.off_day == True: pass

            elif record.action == 'Clock-in':
                try:
                    if last_saved_record[assigned_row].day != record.date.day: 
                        last_saved_record = {}
                        offset = 1
                    if is_cell_empty(assigned_row, assigned_column) and is_date_after_neighbors(record.date, last_saved_record[assigned_row]):
                        assigned_row += offset
                    elif not is_cell_empty(assigned_row, assigned_column):
                        assigned_row += offset
                except KeyError:
                    pass

                populate_cell(assigned_row, assigned_column, record.date.strftime('%H:%M'), font_bold=False)

            else:
                if not is_cell_empty(assigned_row, assigned_column + 1):
                    assigned_row += offset

                populate_cell(assigned_row, assigned_column + 1, record.date.strftime('%H:%M'), font_bold=False)
                populate_cell(assigned_row, assigned_column + 2, record.break_duration, font_bold=False)

            last_saved_record[assigned_row] = record.date
            if records_starting_row != assigned_row:
                offset += 1

        for off_day in records.filter(off_day=True):
            weekday = record.date.weekday() + 1 #TODO properly get date, it comes with 1h less
            assigned_column = records_starting_column + (weekday * 3)
            populate_cell(records_starting_row, assigned_column, "OFF", font_bold=True)
            merge_cells(
                records_starting_row, assigned_column, 
                records_starting_row + offset - 1, assigned_column + 2
            )

        populate_cell(records_starting_row, records_starting_column + 21, worked_hours, border=True)
        merge_cells(
            records_starting_row, records_starting_column + 21, 
            records_starting_row + offset - 1, records_starting_column + 21
        )
        
        users_starting_row = records_starting_row + offset
    


def set_worksheet_styles() :
    global users_starting_row
    global biggest_name_size
    columns = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V']
    for column in columns:
        WORKSHEET.column_dimensions[column].width = 7
    WORKSHEET.column_dimensions["A"].width = biggest_name_size
    WORKSHEET.column_dimensions["W"].width = 10.15
    WORKSHEET.column_dimensions.group(start="X", end="AMJ", hidden=True)

    WORKSHEET.row_dimensions[1].height = 34

    filled_collumns = [2, 3, 4, 8, 9, 10, 14, 15, 16, 20, 21, 22]
    for column in filled_collumns:
        for row in range(4, users_starting_row):
            cell = WORKSHEET.cell(row=row, column=column)
            cell.fill = PatternFill("solid", fgColor=BLUE_COLOR)
    

def merge_cells(start_row, start_column, end_row, end_column):
    WORKSHEET.merge_cells(
        start_row=start_row, start_column=start_column,
        end_row=end_row, end_column=end_column
    )


def populate_cell(row, column, value, color=None, font_size=11, font_bold=True, border=False):
    cell = WORKSHEET.cell(row=row, column=column)
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(name='Arial', size=font_size, bold=font_bold)

    if color:
        cell.fill = PatternFill("solid", fgColor=color)
    if border:
        double = Side(border_style="thick", color="000000")
        cell.border = Border(double, double, double, double)
    

def is_cell_empty(row, column):
    cell = WORKSHEET.cell(row=row, column=column)
    return cell.value == None


def is_date_after_neighbors(date, neighbors_date):
    return date > neighbors_date


def get_week_dates(year, week_number):
    date = f'{year}-W{week_number}'
    first_day = datetime.strptime(date + '-1', "%Y-W%W-%w")
    return [first_day + timedelta(days=d) for d in range(7)]


def ordinal_number(number):
    suffix = None
    match number:
        case 1 | 21 | 31:
            suffix = 'st'
        case 2 | 22:
            suffix = 'nd'
        case 3 | 23:
            suffix = 'rd'
        case _:
            suffix = 'th'
    return f'{number}{suffix}'
